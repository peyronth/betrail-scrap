# import
import requests
import os

# Scrap a file
fileUrl = "http://www.kikourou.net/resultats/resultat-169975-la_course_nature_des_3_etangs_-_18_km-2021.html"
file = requests.get(fileUrl)

# Parse the file
from bs4 import BeautifulSoup
soup = BeautifulSoup(file.content, 'html.parser')

# Get the table
table = soup.find('table', attrs={'id': 'results'})

# Parse the table to get the data
import pandas as pd
df = pd.read_html(str(table))[0]

# Convert column Perf (format %Hh%M'%S'') to seconds
import datetime
df['Perf'] = df['Perf'].apply(lambda x: datetime.datetime.strptime(x, '%Hh%M\'%S\'\'').time())
# df['Perf'] = df['Perf'].apply(lambda x: datetime.timedelta(hours=x.hour, minutes=x.minute, seconds=x.second).total_seconds())

# Create column gender with the last letter of the column Cat
df['Gender'] = df['Cat'].str[-1]

# Remove Cat column
df = df.drop(columns=['Cat'])

# Create an excel file without index and without header
excelFileName = fileUrl.split('/')[-1].split('.')[0] + '.xlsx'
excelFilePath = os.path.join(os.path.dirname(os.path.realpath(__file__)), excelFileName)
df.to_excel(excelFilePath, index=False, header=False)

# change column type of Perf to time
from openpyxl import load_workbook
wb = load_workbook(excelFilePath)
ws = wb.active
for cell in ws['B']:
    cell.number_format = 'hh:mm:ss'
wb.save(excelFilePath)

